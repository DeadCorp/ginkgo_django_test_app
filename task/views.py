import csv
import json
import os
import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect, FileResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from openpyxl import Workbook

from product.models import Product
from supplieraccount.models import SupplierCodes
from task.forms import TaskForm
from task.generate_report_data import ReportData
from task.models import Task


@login_required()
def refresh_product_data(request, sku):
    # create new task with 1 sku for update product data
    obj = Task(user_id=request.user, data=sku)
    obj.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER') + '#' + sku)


@login_required()
def add_task(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            form.save()
        return redirect('task:add_task')
    else:
        form = TaskForm()
        tasks = Task.objects.filter(user_id=request.user).order_by('-id')
    if request.META.get('task_error_id'):
        wrong = {
            'task_error_id': request.META.get('task_error_id'),
            'error': '\n'.join(request.META.get('error')),
        }
    else:
        wrong = ''

    suppliers = SupplierCodes.SUPPLIERS
    return render(request, 'task/add_task.html', {'form': form, 'tasks': tasks, 'suppliers': suppliers, 'wrong': wrong})


@login_required()
def delete_task(request):
    if request.method == 'POST':
        task_id = request.POST.get('task_id')
        try:
            task = Task.objects.get(pk=task_id)
            task.delete()
        except Task.DoesNotExist:
            pass
    return redirect('task:add_task')


@login_required()
def retry_task(request, pk):
    try:
        task = Task.objects.get(pk=pk)
        task.data = ' '.join(json.loads(task.data))
        task.save()
    except Task.DoesNotExist:
        pass
    return redirect(reverse('task:add_task') + '#' + pk)


@login_required()
def task_gen_csv(request, pk):
    if request.method == 'POST':
        data_generator = ReportData(pk, request)

        path_to_file = data_generator.create_file_path('csv')
        with open(path_to_file, mode='w') as file:
            field_names = data_generator.field_names
            writer = csv.DictWriter(file, fieldnames=field_names)
            writer.writeheader()
            for row in data_generator.generate():
                writer.writerow(row)

        if os.path.exists(path_to_file):
            response = FileResponse(open(path_to_file, 'rb'))
            return response
        # return check_generated_errors(data_generator)

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required()
def task_gen_xls(request, pk):
    if request.method == 'POST':
        workbook = Workbook()
        workbook.create_sheet(index=0, title='Sheet1')
        work_sheet = workbook.active

        data_generator = ReportData(pk, request)
        field_names = data_generator.field_names

        cell_range = work_sheet['A1:F1']
        gen = (cell for row in cell_range for cell in row)
        for x, cell in enumerate(gen):
            cell.value = field_names[x]

        path_to_file = data_generator.create_file_path('xls')

        row_num_global = 1
        data = data_generator.generate()

        for product in data:
            gen_cell = (work_sheet.cell(row=row_num + 1, column=col_num + 1) for row_num in range(row_num_global, len(data) + 1)
                        for col_num in range(0, len(field_names)))
            for x, cell in enumerate(gen_cell):
                for i, item in enumerate(product):
                    if i == x:
                        cell.value = product[item]
            row_num_global += 1

        workbook.save(path_to_file)

        if os.path.exists(path_to_file):
            response = FileResponse(open(path_to_file, 'rb'))
            return response
        # return check_generated_errors(data_generator, path_to_file)

    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


def check_generated_errors(data_generator):
    return add_task(data_generator.take_request_with_errors())
