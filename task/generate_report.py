import csv
import datetime
import json
import os
from string import ascii_uppercase

from openpyxl import Workbook

from order.models import OrderStatus
from product.models import Product
from task.models import Task


class ReportData(object):

    def __init__(self, request, task_id=None):
        if task_id:
            self.task_instance = Task.objects.get(pk=task_id)
            self.task_data = json.loads(self.task_instance.data)
            self.wrong = {'task_error_id': self.task_instance.pk,
                          'error': [],
                          }
        self.request = request
        self.request.method = 'GET'
        self.generated_list = []

    @property
    def task_field_names(self):
        return ['SKU', 'Availability', 'Item price', 'Delivery price', 'Total price', 'Quantity']

    @property
    def order_field_names(self):
        return ['Order id', 'SKU', 'Supplier account ID', 'Supplier account email', 'Single price in cart', 'Qty in cart', 'Order status']

    def generate_task_data(self):

        for sku in self.task_data:
            to_write = {
                'SKU': sku,
            }
            try:
                product = Product.objects.get(sku=sku)
                available = product.available if product.delivery_price != 'No delivery' else 'Out of stock'
                to_write.update({
                    'Availability': available,
                })
                if available == 'Out of stock':
                    to_write.update({key: 0 for key in self.task_field_names if key not in to_write.keys()})
                else:
                    delivery_price = product.delivery_price if product.delivery_price != 'Free delivery' else 0.00
                    to_write.update({
                        'Item price': product.price,
                        'Delivery price': delivery_price,
                        'Total price': round(float(product.price) + float(delivery_price), 2),
                        'Quantity': product.available_count if product.available_count != 'is unknown' else 1,
                    })
            except Product.DoesNotExist:
                self.add_error(sku)
                to_write.update({key: 'No product in data base' for key in self.task_field_names if key not in to_write.keys()})
            self.generated_list.append(to_write)

        if len(self.wrong['error']) != 0:
            self.add_error(last=True)

        return self.generated_list

    def generate_order_data(self, orders):

        for order in orders:
            order_status = OrderStatus.objects.get(order=order)
            self.generated_list.append({
                'Order id': order.order_id if order.order_id else 'Currently there is no',
                'SKU': order.product.sku,
                'Supplier account ID': order.account.pk,
                'Supplier account email': order.account.email,
                'Single price in cart': order.price if order.price else 'Currently there is no',
                'Qty in cart': order.quantity if order.quantity else 'Currently there is no',
                'Order status': order_status.status

            })
        return self.generated_list

    def add_error(self, sku=None, last=False):
        self.wrong['error'].append(f'Product with SKU - {sku} dont exist in data base.\n' if not last else
                                   f'Click to Retry task for get this product in data base.')
        self.request.META.update(self.wrong)

    def take_request_with_errors(self):
        return self.request

    def create_file_path(self, report_module, report_target, end):
        """

        :param report_module: module for report example 'celery' or 'scrapy'
        :param report_target: target for report example 'order' or 'task'
        :param end: file end like 'csv' or 'xls'
        :return: created path for save file
        """
        dir_name = os.path.join(os.path.dirname(os.path.dirname(__file__)), f'media/{report_module}_{report_target}_{end}')
        if not os.path.isdir(dir_name):
            os.makedirs(dir_name)
        file_name = f'{report_module}_{report_target}_report_generated_{datetime.datetime.now()}.{end}'
        path_to_file = os.path.join(dir_name, file_name)
        return path_to_file


class ReportFile(object):

    def __init__(self, data, field_names, path_to_file):
        self.data = data
        self.field_names = field_names
        self.path_to_file = path_to_file

    def generate_csv_file(self):
        with open(self.path_to_file, mode='w') as file:
            field_names = self.field_names
            writer = csv.DictWriter(file, fieldnames=field_names)
            writer.writeheader()
            for row in self.data:
                writer.writerow(row)

    def generate_xls_file(self):
        workbook = Workbook()
        workbook.create_sheet(index=0, title='Sheet1')
        work_sheet = workbook.active

        letters = [letter for x, letter in enumerate(ascii_uppercase[:len(self.field_names)])]
        cell_range = work_sheet[f'{letters[0]}1:{letters[-1]}1']

        headers = (cell for row in cell_range for cell in row)
        for x, cell in enumerate(headers):
            cell.value = self.field_names[x]

        row_num_global = 1

        for info in self.data:
            gen_cell = (work_sheet.cell(row=row_num + 1, column=col_num + 1) for row_num in
                        range(row_num_global, len(self.data) + 1)
                        for col_num in range(0, len(self.field_names)))
            for x, cell in enumerate(gen_cell):
                for i, item in enumerate(info):
                    if i == x:
                        cell.value = info[item]
            row_num_global += 1
        workbook.save(self.path_to_file)
